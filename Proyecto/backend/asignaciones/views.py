# asignaciones/views.py

from datetime import datetime
import csv
import io
import re
import unicodedata
from typing import List, Dict, Any

from django.db import transaction
from django.db.models import Case, When, Value, IntegerField, Q, Count
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import load_workbook, Workbook
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from drf_spectacular.utils import extend_schema

from core.permissions import AdminFull_TechReadOnlyPlusActions
from usuarios.models import Usuario

from .models import DireccionAsignada, HistorialAsignacion
from .serializers import (
    DireccionAsignadaSerializer,
    HistorialAsignacionSerializer,
    CsvRowResult,
    CargaCSVSerializer,
    AsignarmeActionSerializer,
    EstadoClienteActionSerializer,   # ← usamos este (con motivo opcional)
    ReagendarActionSerializer,
)

# Notificaciones y logs
from core.models import Notificacion, LogSistema
from core.notify import enviar_notificacion_real


# -------- Helpers de normalización / parsing --------
_HEADER_ALIASES = {
    "rut_cliente": ["rut_cliente", "rut", "rut cliente"],
    "id_vivienda": [
        "id_vivienda_cliente", "id_vivienda", "pcs_cliente",
        "customer_id", "customerid"
    ],
    "direccion": [
        "direccion_cliente", "direccion", "direccion del cliente",
        "dirección del cliente", "direccion_del_cliente",
        "direccion_destinatario", "direccion_cliente_del_destinatario"
    ],
    "comuna": [
        "comuna_cliente", "comuna", "comuna del cliente", "comuna_del_cliente"
    ],
    "zona": [
        "zona_cliente", "customer_zone", "zona"  # ← normalizamos a NORTE/SUR/ORIENTE
    ],
    "marca": ["marca", "brand"],
    "tecnologia": ["tecnologia", "customer_network_type"],
    "encuesta": ["encuesta", "encuesta de origen", "survey_type"],
    "id_qualtrics": ["id_qualtrics", "id_de_respuesta", "record_id"],
    "fecha": ["fecha", "fecha_programada", "fecha_registrada", "transactiondate"],
    "bloque": ["bloque", "bloque_horario", "bloque horario reagendamiento"],
    "asignado_email": ["asignado_email", "correo_tecnico", "tecnico_email", "email_tecnico"],
}

_DATE_FORMATS = (
    "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d",
    "%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y/%m/%d %H:%M:%S",
)

def _norm(s: str) -> str:
    if s is None: return ""
    s = str(s).strip().lower()
    for a, b in {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n"}.items():
        s = s.replace(a, b)
    s = re.sub(r"\s+", " ", s).replace(" ", "_")
    return s

def _s(val) -> str:
    if val is None: return ""
    try: return str(val).strip()
    except Exception: return ""

def _build_header_map(raw_headers: List[str]) -> Dict[str, str]:
    norm_headers = { _norm(h): h for h in raw_headers }
    mapping = {}
    for canon, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            k = _norm(alias)
            if k in norm_headers:
                mapping[canon] = norm_headers[k]
                break
    return mapping

def _parse_date(val):
    if not val: return None
    # Si ya viene datetime (openpyxl), extrae date
    if hasattr(val, "year"):
        try: return val.date() if hasattr(val, "date") else val
        except Exception: pass
    s = _s(val)
    for fmt in _DATE_FORMATS:
        try: return datetime.strptime(s, fmt).date()
        except Exception: pass
    # fallback: toma la parte de fecha antes del espacio
    try:
        only_date = s.split()[0]
        for fmt in ("%Y-%m-%d","%d-%m-%Y","%d/%m/%Y","%Y/%m/%d"):
            try: return datetime.strptime(only_date, fmt).date()
            except Exception: pass
    except Exception:
        pass
    return None

def _normalize_bloque(val: str) -> str | None:
    v = _s(val).lower()
    if not v: return None
    if "10" in v and "13" in v: return "10-13"
    if "14" in v and "18" in v: return "14-18"
    if v in {"10-13", "14-18"}: return v
    return None

def _header_score(names_norm: List[str]) -> int:
    score, have = 0, set(names_norm)
    def any_alias(key): return any(_norm(a) in have for a in _HEADER_ALIASES[key])
    for key in ("direccion", "comuna", "id_vivienda", "rut_cliente", "fecha"):
        if any_alias(key): score += 1
    return score

def _rows_from_csv(f) -> List[Dict[str, Any]]:
    data = f.read().decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(data))
    headers = list(reader.fieldnames or [])
    header_map = _build_header_map(headers)
    return [(header_map, r) for r in reader]

def _rows_from_xlsx(f) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows: return []
    header_idx = 0
    for i in range(min(10, len(all_rows))):
        names = [ _s(x) for x in (all_rows[i] or []) ]
        if _header_score([_norm(n) for n in names]) >= 2:
            header_idx = i; break
    headers = [ _s(x) for x in (all_rows[header_idx] or []) ]
    header_map = _build_header_map(headers)
    rows = []
    for raw in all_rows[header_idx+1:]:
        if not any(raw): continue
        r = { headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers)) }
        rows.append((header_map, r))
    return rows

def _canon_get(row: Dict[str, Any], header_map: Dict[str, str], key: str) -> Any:
    src = header_map.get(key)
    return row.get(src) if src else None

def _norm_ec(s: str) -> str:
    s = (str(s) if s is not None else "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")

def _parse_estado_cliente_from_request(data: Dict[str, Any]) -> str | None:
    raw = (data.get("estado_cliente") or data.get("estado") or data.get("q5") or data.get("accion") or "")
    try:
        num = int(str(raw).strip())
    except Exception:
        num = None
    if num is not None:
        return {1:"autoriza",2:"sin_moradores",3:"rechaza",4:"contingencia",5:"masivo",6:"reagendo"}.get(num)
    key = _norm_ec(raw)
    synonyms = {
        "autoriza":"autoriza",
        "sin_moradores":"sin_moradores","sinmoradores":"sin_moradores",
        "rechaza":"rechaza","contingencia":"contingencia","masivo":"masivo",
        "reagendo":"reagendo","reagenda":"reagendo","reagendada":"reagendo","reagendado":"reagendo",
    }
    return synonyms.get(key)

# ---------- Normalizadores de zona / comuna ----------
def _norm_zona(v: str) -> str | None:
    v = _s(v).upper()
    if not v: return None
    v = v.replace("ZONA METROPOLITANA", "").replace("ZONA", "").strip()
    if v.startswith("NORTE"):   return "NORTE"
    if v.startswith("SUR"):     return "SUR"
    # ORIENTE/CENTRO normaliza a ORIENTE (criterio actual)
    if v.startswith("ORIENTE") or v.startswith("CENTRO"):
        return "ORIENTE"
    return None

_COMUNA_ALIASES = {
    "MACU": "Macul",
    "QNOR": "Quinta Normal",
    "MAIP": "Maipú",
    "PROV": "Providencia",
    "LREI": "La Reina",
    "PENA": "Peñalolén",
    "LFLO": "La Florida",
    "PALT": "Puente Alto",
    "RECO": "Recoleta",
    "CERR": "Cerrillos",
    "SBER": "San Bernardo",
    "NUNO": "Ñuñoa",
    "LCON": "Las Condes",
    "QUIL": "Quilicura",
    "CNCH": "Conchalí",
    "SMGL": "San Miguel",
    "ELBO": "El Bosque",
    "LPRA": "Lo Prado",
    "VITA": "Vitacura",
    "COLI": "Colina",
    "RENC": "Renca",
    "ECEN": "Estación Central",
    "LACI": "La Cisterna",
    "INDE": "Independencia",
    "PACE": "Pedro Aguirre Cerda",
    "LGRA": "La Granja",
    "CNAV": "Cerro Navia",
    "LBAR": "Lo Barnechea",
    "HUEC": "Huechuraba",
    "PUDA": "Pudahuel",
    "LAMP": "Lampa",
    "SJOA": "San Joaquín",
    "SJOS": "San José de Maipo",
    "PHUR": "Padre Hurtado",
    "IMAI": "Isla de Maipo",
    "CTAN": "Calera de Tango",
}

_COMUNA_CANON = {
    "NUNOA": "Ñuñoa",
    "PENALOLEN": "Peñalolén",
    "LAS CONDES": "Las Condes",
    "ESTACION CENTRAL": "Estación Central",
    "SAN MIGUEL": "San Miguel",
}

def _norm_comuna(v: str) -> str:
    raw = _s(v)
    if not raw:
        return ""
    u = raw.strip().upper()
    if u in _COMUNA_ALIASES:
        return _COMUNA_ALIASES[u]
    if u in _COMUNA_CANON:
        return _COMUNA_CANON[u]
    try:
        return raw.strip().title()
    except Exception:
        return raw.strip()


# ---------- núcleo de reagendamiento ----------
def _bloque_label(b):
    return "10:00 a 13:00" if b == "10-13" else ("14:00 a 18:00" if b == "14-18" else (b or "-"))

def _registrar_reagendamiento(asignacion, fecha, bloque, usuario, motivo="REAGENDAMIENTO"):
    old_fecha = asignacion.reagendado_fecha
    old_bloque = asignacion.reagendado_bloque

    # Mantener estado; si estaba PENDIENTE la dejamos ASIGNADA
    new_estado = asignacion.estado
    if new_estado == "PENDIENTE":
        new_estado = "ASIGNADA"

    asignacion.reagendado_fecha = fecha
    asignacion.reagendado_bloque = bloque
    asignacion.estado = new_estado
    asignacion.save(update_fields=["reagendado_fecha", "reagendado_bloque", "estado", "updated_at"])

    # Historial
    accion_val = getattr(HistorialAsignacion.Accion, "REAGENDADA", "REAGENDADA")
    detalles = (
        f"Reagendada: "
        f"{old_fecha or '-'} { _bloque_label(old_bloque) } -> "
        f"{fecha} { _bloque_label(bloque) }"
    )
    HistorialAsignacion.objects.create(
        asignacion=asignacion,
        accion=accion_val,
        detalles=detalles if not motivo else f"{detalles} | Motivo: {motivo}",
        usuario=usuario,
    )

    # Log sensible
    try:
        LogSistema.objects.create(
            usuario=usuario if getattr(usuario, "is_authenticated", False) else None,
            accion=getattr(LogSistema.Accion, "ASSIGN_REAGENDAR", "ASSIGN_REAGENDAR"),
            detalle=f"Reagendó asignación #{asignacion.id} -> {fecha} {bloque}. Motivo: {motivo or '-'}",
        )
    except Exception:
        pass

    # (opcional) registro en tabla Reagendamiento si existe
    try:
        from .models import Reagendamiento
        Reagendamiento.objects.create(
            asignacion=asignacion,
            fecha_anterior=old_fecha,
            bloque_anterior=old_bloque,
            fecha_nueva=fecha,
            bloque_nuevo=bloque,
            motivo=(motivo or "REAGENDAMIENTO"),
            usuario=usuario,
        )
    except Exception:
        pass

    # Notificación email al técnico (si tiene correo)
    dest = getattr(asignacion.asignado_a, "email", "") or ""
    notif = Notificacion.objects.create(
        tipo="reagendo",
        asignacion=asignacion,
        canal=Notificacion.Canal.EMAIL if dest else Notificacion.Canal.NONE,
        destino=dest,
        provider="",
        payload={
            "asignacion_id": asignacion.id,
            "direccion": asignacion.direccion,
            "comuna": asignacion.comuna,
            "reagendado_fecha": str(fecha),
            "reagendado_bloque": bloque,
            "tecnico_id": getattr(usuario, "id", None),
            "tecnico_email": dest,
        },
        status=Notificacion.Estado.QUEUED,
    )
    if dest:
        enviar_notificacion_real(notif)


# ---------------------------------------------------
class DireccionAsignadaViewSet(viewsets.ModelViewSet):
    """
    - Administrador: CRUD total + carga CSV/XLSX + reasignar/desasignar.
    - Técnico: lectura + acciones POST (/asignarme, /estado_cliente, /reagendar).
    """
    queryset = DireccionAsignada.objects.all().order_by("-created_at")
    serializer_class = DireccionAsignadaSerializer
    permission_classes = [IsAuthenticated, AdminFull_TechReadOnlyPlusActions]
    tech_allowed_actions = {"asignarme", "desasignarme", "estado_cliente", "reagendar"}

    filterset_fields = ["estado", "comuna", "zona", "marca", "tecnologia", "encuesta", "asignado_a"]
    search_fields = ["direccion", "comuna", "rut_cliente", "id_vivienda"]
    ordering_fields = ["fecha", "created_at"]

    def get_queryset(self):
        u = self.request.user
        qs = DireccionAsignada.objects.all()

        if self.request.query_params.get("mine") and getattr(u, "rol", None) == "tecnico":
            qs = qs.filter(asignado_a_id=u.id)

        order = self.request.query_params.get("order", "").lower()
        if order == "prioridad":
            prioridad = Case(
                When(reagendado_fecha__isnull=False, then=Value(1)),
                When(asignado_a__isnull=False, then=Value(2)),
                default=Value(3),
                output_field=IntegerField(),
            )
            qs = qs.annotate(prioridad=prioridad).order_by("prioridad", "fecha", "id")
        else:
            qs = qs.order_by("fecha", "id")

        fgte = self.request.query_params.get("fecha__gte")
        flte = self.request.query_params.get("fecha__lte")
        if fgte: qs = qs.filter(fecha__gte=fgte)
        if flte: qs = qs.filter(fecha__lte=flte)
        return qs

    # ---------- ASIGNARME (técnico) ----------
    @extend_schema(request=AsignarmeActionSerializer, responses=DireccionAsignadaSerializer)
    @action(detail=True, methods=["get", "post"], url_path="asignarme", serializer_class=AsignarmeActionSerializer)
    def asignarme(self, request, pk=None):
        obj = self.get_object()
        u = request.user

        if request.method == "GET":
            return Response({
                "ayuda": "Haz POST para asignarte esta visita.",
                "nota": "Este endpoint solo asigna la visita. El reagendamiento se hace en /estado_cliente (reagendo) o en /reagendar.",
                "restricciones": [
                    "Solo técnicos pueden asignarse.",
                    "No se puede asignar una visita VISITADA o CANCELADA.",
                    "Si ya está asignada a otro técnico, devuelve 409.",
                ]
            })

        # POST
        if getattr(u, "rol", None) != "tecnico":
            return Response({"detail": "Solo técnicos pueden asignarse visitas."}, status=403)

        if obj.estado in {"VISITADA", "CANCELADA"}:
            return Response({"detail": "No se puede asignar una visita finalizada/cancelada."}, status=409)

        if obj.asignado_a_id and obj.asignado_a_id != u.id:
            return Response({"detail": "La visita ya está asignada a otro técnico."}, status=409)

        obj.asignado_a_id = u.id
        if obj.estado == "PENDIENTE":
            obj.estado = "ASIGNADA"
        obj.save(update_fields=["asignado_a", "estado", "updated_at"])

        HistorialAsignacion.objects.create(
            asignacion=obj,
            accion=HistorialAsignacion.Accion.ASIGNADA_TECNICO,
            detalles=f"Asignada a {u.email}.",
            usuario=u,
        )
        # Log sensible
        try:
            LogSistema.objects.create(
                usuario=u,
                accion=getattr(LogSistema.Accion, "ASSIGN_SELF", "ASSIGN_SELF"),
                detalle=f"Técnico {u.email} se asignó la visita #{obj.id}",
            )
        except Exception:
            pass

        # responder con la dirección completa
        return Response(DireccionAsignadaSerializer(obj, context={"request": request}).data)

    # ---------- DESASIGNARME (técnico) ----------
    @action(detail=True, methods=["post"], url_path="desasignarme")
    def desasignarme(self, request, pk=None):
        """
        Permite al TÉCNICO quitarse una visita que está asignada a él,
        siempre que la visita no esté VISITADA o CANCELADA.
        No toca reagendos existentes (solo suelta la asignación).
        """
        obj = self.get_object()
        u = request.user

        # Solo técnicos
        if getattr(u, "rol", None) != "tecnico":
            return Response({"detail": "Solo técnicos pueden desasignarse."}, status=403)

        # Debe estar asignada a este técnico
        if obj.asignado_a_id != u.id:
            return Response({"detail": "Solo puedes desasignarte de tus propias visitas."}, status=409)

        # No permitir si ya está finalizada/cancelada
        if obj.estado in {"VISITADA", "CANCELADA"}:
            return Response({"detail": "No se puede desasignar una visita finalizada/cancelada."}, status=409)

        motivo = (request.data or {}).get("motivo", "").strip()

        with transaction.atomic():
            obj.asignado_a = None
            # Si estaba ASIGNADA, vuelve a PENDIENTE (criterio actual)
            obj.estado = "PENDIENTE"
            obj.save(update_fields=["asignado_a", "estado", "updated_at"])

            HistorialAsignacion.objects.create(
                asignacion=obj,
                accion=HistorialAsignacion.Accion.DESASIGNADA,
                detalles="Desasignada por el propio técnico." + (f" Motivo: {motivo}" if motivo else ""),
                usuario=u,
            )

        # Log sensible
        try:
            LogSistema.objects.create(
                usuario=u,
                accion=getattr(LogSistema.Accion, "ASSIGN_UNASSIGN_SELF", "ASSIGN_UNASSIGN_SELF"),
                detalle=f"Técnico {u.email} se desasignó la visita #{obj.id}. Motivo: {motivo or '-'}",
            )
        except Exception:
            pass

        return Response(DireccionAsignadaSerializer(obj, context={"request": request}).data)

    # ---------- HISTORIAL ----------
    @action(detail=False, methods=["get"], url_path="historial")
    def historial(self, request):
        u = request.user
        tecnico_id = request.query_params.get("tecnico_id")
        qs = HistorialAsignacion.objects.select_related("asignacion", "usuario").all()

        if getattr(u, "rol", None) == "tecnico":
            qs = qs.filter(Q(usuario_id=u.id) | Q(asignacion__asignado_a_id=u.id))
        else:
            if tecnico_id:
                qs = qs.filter(Q(usuario_id=tecnico_id) | Q(asignacion__asignado_a_id=tecnico_id))

        estado = request.query_params.get("estado")
        marca = request.query_params.get("marca")
        tecnologia = request.query_params.get("tecnologia")
        comuna = request.query_params.get("comuna")
        zona = request.query_params.get("zona")
        encuesta = request.query_params.get("encuesta")

        if estado: qs = qs.filter(asignacion__estado=estado)
        if marca: qs = qs.filter(asignacion__marca=marca)
        if tecnologia: qs = qs.filter(asignacion__tecnologia=tecnologia)
        if comuna: qs = qs.filter(asignacion__comuna=comuna)
        if zona: qs = qs.filter(asignacion__zona=zona)
        if encuesta: qs = qs.filter(asignacion__encuesta=encuesta)

        desde = request.query_params.get("fecha_after") or request.query_params.get("desde")
        hasta = request.query_params.get("fecha_before") or request.query_params.get("hasta")
        hoy = timezone.localdate().isoformat()
        if desde and str(desde).upper() == "HOY": desde = hoy
        if hasta and str(hasta).upper() == "HOY": hasta = hoy
        if desde: qs = qs.filter(asignacion__fecha__gte=desde)
        if hasta: qs = qs.filter(asignacion__fecha__lte=hasta)

        qs = qs.order_by("-created_at", "-id")
        page = self.paginate_queryset(qs)
        s = HistorialAsignacionSerializer(page or qs, many=True)
        return self.get_paginated_response(s.data) if page is not None else Response(s.data)

    @action(detail=False, methods=["post"], url_path="historial/export")
    def historial_export(self, request):
        u = request.user
        if getattr(u, "rol", None) != "administrador":
            return Response({"detail": "Solo administrador puede exportar historial."}, status=403)

        qs = HistorialAsignacion.objects.select_related("asignacion", "usuario").all()
        p = self.request.query_params
        if p.get("estado"): qs = qs.filter(asignacion__estado=p["estado"])
        if p.get("marca"): qs = qs.filter(asignacion__marca=p["marca"])
        if p.get("tecnologia"): qs = qs.filter(asignacion__tecnologia=p["tecnologia"])
        if p.get("comuna"): qs = qs.filter(asignacion__comuna=p["comuna"])
        if p.get("zona"): qs = qs.filter(asignacion__zona=p["zona"])
        if p.get("encuesta"): qs = qs.filter(asignacion__encuesta=p["encuesta"])

        desde = p.get("fecha_after") or p.get("desde")
        hasta = p.get("fecha_before") or p.get("hasta")
        hoy = timezone.localdate().isoformat()
        if desde and str(desde).upper() == "HOY": desde = hoy
        if hasta and str(hasta).upper() == "HOY": hasta = hoy
        if desde: qs = qs.filter(asignacion__fecha__gte=desde)
        if hasta: qs = qs.filter(asignacion__fecha__lte=hasta)

        fmt = (request.data or {}).get("format", "xlsx").lower()
        qs = qs.order_by("asignacion_id", "created_at", "id")

        if fmt == "csv":
            buff = io.StringIO()
            writer = csv.writer(buff)
            writer.writerow(["historial_id","asignacion_id","direccion","comuna","marca","tecnologia","fecha","bloque","accion","detalles","usuario_email","creado"])
            for h in qs.iterator():
                writer.writerow([
                    h.id, h.asignacion_id,
                    getattr(h.asignacion, "direccion", ""),
                    getattr(h.asignacion, "comuna", ""),
                    getattr(h.asignacion, "marca", ""),
                    getattr(h.asignacion, "tecnologia", ""),
                    getattr(h.asignacion, "fecha", "") or "",
                    getattr(h.asignacion, "reagendado_bloque", "") or "",
                    h.accion, h.detalles,
                    getattr(h.usuario, "email", ""),
                    h.created_at.isoformat(),
                ])
            resp = HttpResponse(buff.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="historial.csv"'
            return resp

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        ws.append(["historial_id","asignacion_id","direccion","comuna","marca","tecnologia","fecha","bloque","accion","detalles","usuario_email","creado"])
        for h in qs.iterator():
            ws.append([
                h.id, h.asignacion_id,
                getattr(h.asignacion, "direccion", ""),
                getattr(h.asignacion, "comuna", ""),
                getattr(h.asignacion, "marca", ""),
                getattr(h.asignacion, "tecnologia", ""),
                getattr(h.asignacion, "fecha", "") or "",
                getattr(h.asignacion, "reagendado_bloque", "") or "",
                h.accion, h.detalles,
                getattr(h.usuario, "email", ""),
                h.created_at.isoformat(),
            ])
        out = io.BytesIO()
        wb.save(out)
        resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="historial.xlsx"'
        return resp

    # === CARGA CSV/XLSX ===
    @extend_schema(request=CargaCSVSerializer)
    @action(detail=False, methods=["post"], url_path="cargar_csv",
            parser_classes=[MultiPartParser, FormParser], serializer_class=CargaCSVSerializer)
    def cargar_csv(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Sube un archivo en el campo 'file'."}, status=400)

        name = (_s(f.name)).lower()
        try:
            rows = _rows_from_xlsx(f) if name.endswith(".xlsx") else _rows_from_csv(f)
        except Exception as e:
            return Response({"detail": f"No se pudo leer el archivo: {e}"}, status=400)

        results = []
        today = timezone.localdate()

        for idx, (header_map, row) in enumerate(rows, start=2):
            errors, created, updated = [], False, False

            rut_cliente   = _s(_canon_get(row, header_map, "rut_cliente"))
            id_vivienda   = _s(_canon_get(row, header_map, "id_vivienda"))
            direccion     = _s(_canon_get(row, header_map, "direccion"))
            comuna_raw    = _s(_canon_get(row, header_map, "comuna"))
            comuna        = _norm_comuna(comuna_raw)
            zona          = _norm_zona(_canon_get(row, header_map, "zona"))

            marca      = (_s(_canon_get(row, header_map, "marca")) or "CLARO").upper()
            tecnologia = (_s(_canon_get(row, header_map, "tecnologia")) or "HFC").upper()
            encuesta   = (_s(_canon_get(row, header_map, "encuesta")) or "post_visita").lower()
            id_qual    = _s(_canon_get(row, header_map, "id_qualtrics"))

            fecha_val  = _parse_date(_canon_get(row, header_map, "fecha"))
            bloque     = _normalize_bloque(_canon_get(row, header_map, "bloque"))
            asignado_email = _s(_canon_get(row, header_map, "asignado_email"))

            if fecha_val and fecha_val < today:
                fecha_val = None

            if not direccion or not comuna:
                errors.append("Faltan campos obligatorios: direccion/comuna.")

            asignado = None
            if asignado_email:
                asignado = Usuario.objects.filter(email__iexact=asignado_email, rol="tecnico").first()
                if not asignado:
                    errors.append("asignado_email no existe o no es técnico.")

            if errors:
                results.append({"rownum": idx, "created": False, "updated": False, "errors": errors})
                continue

            defaults = {
                "rut_cliente": rut_cliente,
                "direccion": direccion,
                "comuna": comuna,           # ← normalizada
                "marca": marca or "CLARO",
                "tecnologia": tecnologia or "HFC",
                "encuesta": encuesta or "post_visita",
                "id_qualtrics": id_qual,
            }
            if fecha_val:
                defaults["fecha"] = fecha_val
            if zona:
                defaults["zona"] = zona

            if id_vivienda:
                obj, was_created = DireccionAsignada.objects.get_or_create(
                    id_vivienda=id_vivienda,
                    defaults={**defaults, "estado": "PENDIENTE"},
                )
            else:
                obj, was_created = DireccionAsignada.objects.get_or_create(
                    direccion=direccion,
                    comuna=comuna,
                    defaults={**defaults, "estado": "PENDIENTE"},
                )

            for k, v in defaults.items():
                setattr(obj, k, v)

            # No importar "bloque" como reagendo (criterio actual)
            if bloque:
                obj.reagendado_bloque = None

            if asignado:
                obj.asignado_a = asignado
                if not obj.fecha and fecha_val:
                    obj.fecha = fecha_val
                obj.estado = "ASIGNADA"
            else:
                obj.asignado_a = None
                obj.estado = "PENDIENTE"

            obj.save()

            if was_created:
                HistorialAsignacion.objects.create(
                    asignacion=obj,
                    accion=HistorialAsignacion.Accion.CREADA,
                    detalles=f"Creada por carga XLS/CSV. {'Asignada a ' + asignado.email if asignado else 'Sin técnico'}",
                    usuario=request.user,
                )
                created = True
            else:
                HistorialAsignacion.objects.create(
                    asignacion=obj,
                    accion=HistorialAsignacion.Accion.EDITADA,  # ← ahora existe
                    detalles="Actualizada por carga XLS/CSV.",
                    usuario=request.user,
                )
                updated = True

            results.append({"rownum": idx, "created": created, "updated": updated})

        # Log resumen de carga
        try:
            LogSistema.objects.create(
                usuario=request.user if getattr(request.user, "is_authenticated", False) else None,
                accion=getattr(LogSistema.Accion, "ASSIGN_BULK_LOAD", "ASSIGN_BULK_LOAD"),
                detalle=f"Carga {('XLSX' if name.endswith('.xlsx') else 'CSV')} '{_s(getattr(f, 'name', 'desconocido'))}': "
                        f"creados={sum(1 for r in results if r['created'])}, "
                        f"actualizados={sum(1 for r in results if r['updated'])}, "
                        f"errores={sum(1 for r in results if r.get('errors'))}",
            )
        except Exception:
            pass

        return Response({
            "ok": True,
            "rows": results,
            "summary": {
                "created": sum(1 for r in results if r["created"]),
                "updated": sum(1 for r in results if r["updated"]),
                "errors":  sum(1 for r in results if r.get("errors")),
            }
        })

    # ---------- DESASIGNAR (ADMIN) ----------
    @action(detail=True, methods=["patch"], url_path="desasignar")
    def desasignar(self, request, pk=None):
        obj = self.get_object()
        if getattr(request.user, "rol", None) != "administrador":
            return Response({"detail": "Solo administrador puede desasignar."}, status=403)
        obj.asignado_a = None
        obj.estado = "PENDIENTE"
        obj.save()
        HistorialAsignacion.objects.create(
            asignacion=obj,
            accion=HistorialAsignacion.Accion.DESASIGNADA,  # ← existe en tu modelo
            detalles="Desasignada por administrador.",
            usuario=request.user,
        )
        # Log sensible
        try:
            LogSistema.objects.create(
                usuario=request.user,
                accion=getattr(LogSistema.Accion, "ASSIGN_UNASSIGN_ADMIN", "ASSIGN_UNASSIGN_ADMIN"),
                detalle=f"Admin {request.user.email} desasignó la visita #{obj.id}",
            )
        except Exception:
            pass
        return Response(DireccionAsignadaSerializer(obj, context={"request": request}).data)

    # ---------- ESTADO DEL CLIENTE (Q5) ----------
    @extend_schema(request=EstadoClienteActionSerializer, responses=DireccionAsignadaSerializer)
    @action(detail=True, methods=["get", "post"], url_path="estado_cliente", serializer_class=EstadoClienteActionSerializer)
    def estado_cliente(self, request, pk=None):
        obj = self.get_object()

        if request.method == "GET":
            return Response({
                "ayuda": "POST con {'estado_cliente': ..., opcionalmente 'reagendado_fecha','reagendado_bloque','motivo'}. "
                         "Si eliges 'Reagendó', debes indicar fecha/bloque.",
            })

        ser = EstadoClienteActionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        estado = ser.validated_data["estado_cliente"]
        motivo = (ser.validated_data.get("motivo") or "").strip()

        with transaction.atomic():
            if estado == "autoriza":
                obj.estado = "VISITADA"
                obj.save(update_fields=["estado", "updated_at"])
                HistorialAsignacion.objects.create(
                    asignacion=obj,
                    accion=HistorialAsignacion.Accion.ESTADO_CLIENTE,
                    detalles="Estado cliente = autoriza (asignación VISITADA).",
                    usuario=request.user,
                )
                # Log
                try:
                    LogSistema.objects.create(
                        usuario=request.user,
                        accion=getattr(LogSistema.Accion, "ASSIGN_MARK_VISITED", "ASSIGN_MARK_VISITED"),
                        detalle=f"Marcó VISITADA la asignación #{obj.id}",
                    )
                except Exception:
                    pass

            elif estado in {"sin_moradores", "rechaza", "contingencia", "masivo"}:
                obj.estado = "CANCELADA"
                obj.save(update_fields=["estado", "updated_at"])
                HistorialAsignacion.objects.create(
                    asignacion=obj,
                    accion=HistorialAsignacion.Accion.ESTADO_CLIENTE,
                    detalles=f"Estado cliente = {estado} (asignación CANCELADA).",
                    usuario=request.user,
                )
                # Log
                try:
                    LogSistema.objects.create(
                        usuario=request.user,
                        accion=getattr(LogSistema.Accion, "ASSIGN_MARK_CANCELLED", "ASSIGN_MARK_CANCELLED"),
                        detalle=f"Marcó CANCELADA la asignación #{obj.id} (motivo cliente={estado})",
                    )
                except Exception:
                    pass

            else:  # reagendo
                f = ser.validated_data["reagendado_fecha"]     # date validado
                b = ser.validated_data["reagendado_bloque"]    # "10-13"/"14-18"
                motivo_final = motivo or "Q5(estado_cliente=reagendo)"
                _registrar_reagendamiento(obj, f, b, request.user, motivo=motivo_final)

        serializer = DireccionAsignadaSerializer(obj, context={"request": request})
        return Response(serializer.data)

    # ---------- REAGENDAR (endpoint dedicado) ----------
    @extend_schema(request=ReagendarActionSerializer, responses=DireccionAsignadaSerializer)
    @action(detail=True, methods=["get", "post"], url_path="reagendar", serializer_class=ReagendarActionSerializer)
    def reagendar(self, request, pk=None):
        obj = self.get_object()

        if request.method == "GET":
            return Response({"ayuda": "POST {fecha:'YYYY-MM-DD', bloque:'10-13|14-18', motivo?} para reagendar."})

        ser = ReagendarActionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        fecha = ser.validated_data["fecha"]
        bloque = ser.validated_data["bloque"]
        # motivo opcional (aunque el serializer no lo defina, lo tomamos del request)
        motivo = _s((request.data or {}).get("motivo"))

        if obj.estado in {"VISITADA", "CANCELADA"}:
            return Response({"detail": "No se puede reagendar una visita finalizada/cancelada."}, status=409)

        with transaction.atomic():
            _registrar_reagendamiento(obj, fecha, bloque, request.user, motivo=motivo or "REAGENDAR")

        return Response(DireccionAsignadaSerializer(obj, context={"request": request}).data)

    # ---------- MÉTRICAS ----------
    @action(detail=False, methods=["get"], url_path="metrics/resumen")
    def metrics_resumen(self, request):
        qs = DireccionAsignada.objects.all()
        p = request.query_params

        if p.get("fecha__gte"): qs = qs.filter(fecha__gte=p["fecha__gte"])
        if p.get("fecha__lte"): qs = qs.filter(fecha__lte=p["fecha__lte"])
        if p.get("tecnico_id"): qs = qs.filter(asignado_a_id=p["tecnico_id"])
        if p.get("zona"): qs = qs.filter(zona=p["zona"])
        if p.get("comuna"): qs = qs.filter(comuna=p["comuna"])
        if p.get("marca"): qs = qs.filter(marca=p["marca"])
        if p.get("tecnologia"): qs = qs.filter(tecnologia=p["tecnologia"])

        total = qs.count()
        c = lambda s: qs.filter(estado=s).count()
        pend, asig, vis, canc = (c("PENDIENTE"), c("ASIGNADA"), c("VISITADA"), c("CANCELADA"))
        # Opción A: reagendadas por existencia de reagendado_fecha
        reag = qs.filter(reagendado_fecha__isnull=False).count()

        pct = lambda n: round(100.0 * n / total, 1) if total else 0.0
        return Response({
            "total": total,
            "pendientes": pend,
            "asignadas": asig,
            "visitadas": vis,
            "canceladas": canc,
            "reagendadas": reag,
            "pct_visitadas": pct(vis),
            "pct_reagendadas": pct(reag),
        })

    @action(detail=False, methods=["get"], url_path="metrics/tecnico")
    def metrics_tecnico(self, request):
        qs = DireccionAsignada.objects.all()
        p = request.query_params
        u = request.user

        tecnico_id = p.get("tecnico_id")
        if getattr(u, "rol", None) == "tecnico":
            tecnico_id = u.id
        if tecnico_id:
            qs = qs.filter(asignado_a_id=tecnico_id)

        if p.get("fecha__gte"): qs = qs.filter(fecha__gte=p["fecha__gte"])
        if p.get("fecha__lte"): qs = qs.filter(fecha__lte=p["fecha__lte"])
        if p.get("zona"): qs = qs.filter(zona=p["zona"])
        if p.get("comuna"): qs = qs.filter(comuna=p["comuna"])
        if p.get("marca"): qs = qs.filter(marca=p["marca"])
        if p.get("tecnologia"): qs = qs.filter(tecnologia=p["tecnologia"])

        total = qs.count()
        c = lambda s: qs.filter(estado=s).count()
        pend, asig, vis, canc = (c("PENDIENTE"), c("ASIGNADA"), c("VISITADA"), c("CANCELADA"))
        # Opción A: reagendadas por existencia de reagendado_fecha
        reag = qs.filter(reagendado_fecha__isnull=False).count()

        pct = lambda n: round(100.0 * n / total, 1) if total else 0.0
        return Response({
            "total": total,
            "pendientes": pend,
            "asignadas": asig,
            "visitadas": vis,
            "canceladas": canc,
            "reagendadas": reag,
            "pct_visitadas": pct(vis),
            "pct_reagendadas": pct(reag),
        })

    @action(detail=False, methods=["post"], url_path="metrics/export")
    def metrics_export(self, request):
        if getattr(request.user, "rol", None) != "administrador":
            return Response({"detail": "Solo administrador puede exportar métricas."}, status=403)

        resumen = self.metrics_resumen(request).data
        fmt = (request.data or {}).get("format", "xlsx").lower()

        if fmt == "csv":
            buff = io.StringIO()
            w = csv.writer(buff)
            w.writerow(["total","pendientes","asignadas","visitadas","canceladas","reagendadas","pct_visitadas","pct_reagendadas"])
            w.writerow([
                resumen["total"], resumen["pendientes"], resumen["asignadas"],
                resumen["visitadas"], resumen["canceladas"], resumen["reagendadas"],
                resumen["pct_visitadas"], resumen["pct_reagendadas"]
            ])
            resp = HttpResponse(buff.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="metricas_resumen.csv"'
            return resp

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["total","pendientes","asignadas","visitadas","canceladas","reagendadas","pct_visitadas","pct_reagendadas"])
        ws.append([
            resumen["total"], resumen["pendientes"], resumen["asignadas"],
            resumen["visitadas"], resumen["canceladas"], resumen["reagendadas"],
            resumen["pct_visitadas"], resumen["pct_reagendadas"]
        ])
        out = io.BytesIO()
        wb.save(out)
        resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="metricas_resumen.xlsx"'
        return resp

    @action(detail=False, methods=["get"], url_path="metrics/serie")
    def metrics_serie(self, request):
        qs = DireccionAsignada.objects.all()
        p = request.query_params

        if p.get("fecha__gte"): qs = qs.filter(fecha__gte=p["fecha__gte"])
        if p.get("fecha__lte"): qs = qs.filter(fecha__lte=p["fecha__lte"])
        if p.get("tecnico_id"): qs = qs.filter(asignado_a_id=p["tecnico_id"])
        if p.get("zona"): qs = qs.filter(zona=p["zona"])
        if p.get("comuna"): qs = qs.filter(comuna=p["comuna"])
        if p.get("marca"): qs = qs.filter(marca=p["marca"])
        if p.get("tecnologia"): qs = qs.filter(tecnologia=p["tecnologia"])

        def serie_por_estado(estado: str):
            base = qs.filter(estado=estado, fecha__isnull=False)
            agg = (base.annotate(d=TruncDate("fecha"))
                        .values("d").annotate(c=Count("id")).order_by("d"))
            return {a["d"].isoformat(): a["c"] for a in agg}

        # visitadas por fecha de visita (fecha)
        v = serie_por_estado("VISITADA")

        # Opción A: reagendadas por fecha de reagendo (reagendado_fecha)
        base_r = qs.filter(reagendado_fecha__isnull=False)
        agg_r = (base_r.annotate(d=TruncDate("reagendado_fecha"))
                        .values("d").annotate(c=Count("id")).order_by("d"))
        r = {a["d"].isoformat(): a["c"] for a in agg_r}

        # canceladas por fecha de cita (fecha)
        c = serie_por_estado("CANCELADA")

        fechas = sorted(set(v.keys()) | set(r.keys()) | set(c.keys()))
        data = {
            "labels": fechas,
            "series": {
                "visitadas": [v.get(d, 0) for d in fechas],
                "reagendadas": [r.get(d, 0) for d in fechas],
                "canceladas": [c.get(d, 0) for d in fechas],
            }
        }
        return Response(data)
